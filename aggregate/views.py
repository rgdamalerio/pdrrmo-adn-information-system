from aggregate.models import AggregatedFamiliesandPopulation
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.db.models import Sum


def index(request): 
    return render(request,'aggregate/index.html')


def exportFamilyandPopulation(request):
    try:
        user = request.user
        # Check if user belongs to the "municipality" group or is an admin
        if user.groups.filter(name='municipality').exists() or user.is_superuser:
            if user.is_superuser:
                aggregated = AggregatedFamiliesandPopulation.objects.all()
            else:
                user_location = user.userlocation
                municipality = user_location.psgccode_mun
                aggregated = AggregatedFamiliesandPopulation.objects.filter(munname=municipality)

            workbook = Workbook()
            worksheet = workbook.active

            worksheet.print_options.fit_to_page = True
            worksheet.print_options.fit_to_width = 1

            bold_font = Font(bold=True)
            fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')
            headers = [
                'Municipality', 'Barangay', 'No. of Households', 'No. of Families', 'Individuals (M)', 'Individuals (F)',
                'Infant 0-11months (M)', 'Infant 0-11months (F)', 'Children 1-17y/o (M)', 'Children 1-17y/o (F)',
                'Adult 18-59y/o (M)', 'Adult 18-59y/o (F)', 'Elderly 60y/o above (M)', 'Elderly 60y/o above (F)',
                'IP (M)', 'IP (F)'
            ]

            for col_num, header_title in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header_title)
                cell.font = bold_font
                cell.fill = fill
                column_letter = get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = 18

            for data in aggregated:
                worksheet.append([
                    data.munname, data.brgyname, data.households, data.families, data.male, data.female, data.male_infant,
                    data.female_infant, data.male_children, data.female_children, data.male_adult,
                    data.female_adult, data.male_elderly, data.female_elderly, data.ip_male, data.ip_female
                ])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=FamiliesandPopulation.xlsx'
            workbook.save(response)
            return response

        else:
            return HttpResponse('You do not have permission to export data')

    except Exception as e:
        error_message = f'An error occurred: {str(e)}'
        return render(request, 'aggregate/error_template.html', {'error_message': error_message})


def chart_view(request):
    return render(request, 'aggregate/population_chart.html')

def generate_chart_data():
    age_groups = [
        ('Infant 0-11months', 'male_infant', 'female_infant'),
        ('Children 1-17y/o', 'male_children', 'female_children'),
        ('Adult 18-59y/o', 'male_adult', 'female_adult'),
        ('Elderly 60y/o above', 'male_elderly', 'female_elderly'),
    ]

    age_group_population = []
    for age_group, male_field, female_field in age_groups:
        total_population = AggregatedFamiliesandPopulation.objects.aggregate(
            total_male=Sum(male_field),
            total_female=Sum(female_field)
        )
        total_population = total_population['total_male'] + total_population['total_female']

        age_group_population.append({
            'Age Group': age_group,
            'Total Population': total_population
        })

    infant_counts = AggregatedFamiliesandPopulation.objects.values('munname').annotate(
        total_infants=Sum('male_infant') + Sum('female_infant')
    )

    chart_data = {
        'age_group_population': age_group_population,
        'infant_counts': []
    }
    for result in infant_counts:
        municipality_name = result['munname']
        total_infants = result['total_infants']
        chart_data['infant_counts'].append({'Municipality': municipality_name, 'Total Infants': total_infants})

    return chart_data

def get_chart_data(request):
    chart_data = generate_chart_data()
    return JsonResponse(chart_data, safe=False)


def infant_counts_report(request):
    return render(request, 'chart.html')
